#!/usr/bin/env python3
import csv, io, json, re, statistics
from datetime import datetime, timezone
from pathlib import Path
import requests
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
PRICES = ROOT / 'data' / 'prices.json'
MIMIT_URLS = ['https://www.mimit.gov.it/images/exportCSV/prezzo_alle_8.csv','https://www.mise.gov.it/images/exportCSV/prezzo_alle_8.csv']
EU_PAGE = 'https://energy.ec.europa.eu/data-and-analysis/weekly-oil-bulletin_en'
KEYS = {'benzina':['benzina'], 'gasolio':['gasolio','diesel'], 'gpl':['gpl'], 'metano':['metano']}
UNITS = {'benzina':'eur_l','gasolio':'eur_l','gpl':'eur_l','metano':'eur_kg'}

def load(p): return json.loads(p.read_text(encoding='utf-8')) if p.exists() else {}
def save(p,o): p.write_text(json.dumps(o, ensure_ascii=False, indent=2)+'\n', encoding='utf-8')
def cell(x): return str(x or '').replace('\ufeff','').strip()
def norm(x): return cell(x).lower().replace(' ','').replace('_','')
def val(x):
    try: return float(cell(x).replace(',','.'))
    except Exception: return None

def find_col(cols, wants):
    for c in cols:
        n = norm(c)
        if any(norm(w) in n for w in wants): return c
    return None

def rows_after_header(text):
    try: dialect = csv.Sniffer().sniff(text[:5000], delimiters=';,|\t')
    except Exception:
        dialect = csv.excel; dialect.delimiter=';'
    rows = [[cell(c) for c in r] for r in csv.reader(io.StringIO(text), dialect=dialect)]
    rows = [r for r in rows if any(r)]
    for i,r in enumerate(rows):
        nr = [norm(c) for c in r]
        if any('carburante' in c or 'prodotto' in c for c in nr) and any('prezzo' in c for c in nr):
            out=[]
            for rr in rows[i+1:]:
                rr = (rr + ['']*len(r))[:len(r)]
                out.append(dict(zip(r, rr)))
            return r,out,i
    raise ValueError('header not found: '+repr(rows[:4]))

def fuel_key(label):
    t = str(label or '').lower()
    if any(x in t for x in ['special','premium','plus']): return None
    for k, aliases in KEYS.items():
        if any(a in t for a in aliases): return k
    return None

def mimit():
    for url in MIMIT_URLS:
        try:
            r = requests.get(url, timeout=(6,20), headers={'User-Agent':'Mozilla/5.0 elettrica-tco'})
            r.raise_for_status()
            cols, rows, skipped = rows_after_header(r.content.decode('utf-8-sig', errors='replace'))
            fc, pc, sc = find_col(cols,['descCarburante','carburante','prodotto']), find_col(cols,['prezzo','price']), find_col(cols,['isSelf','self'])
            buckets = {k:[] for k in KEYS}
            for row in rows:
                k, price = fuel_key(row.get(fc,'')), val(row.get(pc,''))
                if not k or not price or price <= 0 or price > 10: continue
                if sc and str(row.get(sc,'')).lower() in ['0','false','servito'] and k in ['benzina','gasolio']: continue
                buckets[k].append(price)
            data = {k:round(statistics.fmean(v),3) for k,v in buckets.items() if v}
            print('MIMIT parsed', {'url':url,'skipped_rows':skipped,'samples':{k:len(v) for k,v in buckets.items()}})
            if data.get('benzina') and data.get('gasolio'):
                data.update({'source':url,'frequency':'daily','samples':{k:len(v) for k,v in buckets.items()}})
                return data
        except Exception as e:
            print('MIMIT source failed', url, e)
    return {}

def eu_weekly():
    try:
        page = requests.get(EU_PAGE, timeout=(6,20), headers={'User-Agent':'Mozilla/5.0 elettrica-tco'})
        page.raise_for_status()
        links = re.findall(r'href="([^"]+?\.xlsx[^"]*)"', page.text)
        url = None
        for l in links:
            if any(x in l.lower() for x in ['weekly','tax','price']):
                url = ('https://energy.ec.europa.eu'+l if l.startswith('/') else l).replace('&amp;','&'); break
        if not url: return {}
        x = requests.get(url, timeout=(6,35), headers={'User-Agent':'Mozilla/5.0 elettrica-tco'}); x.raise_for_status()
        wb = load_workbook(io.BytesIO(x.content), data_only=True, read_only=True)
        candidates=[]; italy=0
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                cells=[cell(v) for v in row]
                if any(c.lower() in ['italy','italia','it'] for c in cells):
                    italy += 1
                    nums=[val(c) for c in cells]; nums=[n for n in nums if n and .5 <= n <= 3]
                    if len(nums) >= 2: candidates.append(nums)
        if not candidates: return {}
        nums=candidates[0]
        data={'benzina':round(max(nums[:3]),3),'gasolio':round(sorted(nums[:4])[1] if len(nums)>=4 else nums[1],3),'source':url,'frequency':'weekly_eu','samples':{'italy_rows':italy}}
        print('EU weekly parsed', data)
        return data
    except Exception as e:
        print('EU source failed', e); return {}

def main():
    prices = load(PRICES) or {'fuel':{}, 'electricity':{'home':.30,'solar':.08}}
    prices.setdefault('fuel', {}).setdefault('units', UNITS)
    data = mimit(); status='updated_mimit_daily'
    if not data:
        data = eu_weekly(); status='updated_eu_weekly_partial_keep_gpl_metano_previous'
    if data:
        for k in KEYS:
            if data.get(k): prices['fuel'][k]=data[k]
        prices['fuel'].update({'source':data.get('source'), 'frequency':data.get('frequency'), 'samples':data.get('samples',{}), 'units':UNITS})
        prices['status']=status
    else:
        prices['fuel']['units']=UNITS; prices['status']='fallback_previous_values'
    prices['updated_at']=datetime.now(timezone.utc).isoformat(); save(PRICES, prices); print(json.dumps(prices, ensure_ascii=False, indent=2))
if __name__ == '__main__': main()
