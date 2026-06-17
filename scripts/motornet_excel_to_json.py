#!/usr/bin/env python3
"""Convert a cleaned Motornet Excel workbook back to cars_motornet.json.

Default behavior is conservative:
- Load the original JSON as base.
- Match Excel rows by `id`.
- Update only curated fields exported by audit_motornet_quality.py.
- Ignore empty Excel cells unless --clear-empty is passed.
- Preserve all raw/debug fields not present in the Excel file, including specs_raw.
"""
from __future__ import annotations

import argparse
import copy
import json
import math
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


# Keep this list aligned with audit_motornet_quality.py.
CURATED_COLUMNS = [
    "issues",
    "status",
    "notes",
    "id",
    "category",
    "brand",
    "model",
    "version",
    "powertrain",
    "fuel",
    "year",
    "price_eur",
    "price_source",
    "power_kw",
    "power_cv",
    "consumption_l_100km",
    "consumption_kg_100km",
    "consumption_kwh_100km",
    "battery_kwh",
    "range_wltp_km",
    "emissions_g_km",
    "image_url",
    "source_url",
    "motornet_detail_url",
]

TEXT_FIELDS = {
    "category",
    "brand",
    "model",
    "version",
    "powertrain",
    "fuel",
    "price_source",
    "image_url",
    "source_url",
    "motornet_detail_url",
}

NUMERIC_FIELDS = {
    "year",
    "price_eur",
    "power_kw",
    "power_cv",
    "consumption_l_100km",
    "consumption_kg_100km",
    "consumption_kwh_100km",
    "battery_kwh",
    "range_wltp_km",
    "emissions_g_km",
}

IGNORED_FIELDS = {
    "issues",
    "status",
    "notes",
    "id",
}


def clean(value: Any) -> str:
    return " ".join(str(value or "").split())


def parse_number(value: Any) -> float | int | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        n = float(value)
        if not math.isfinite(n):
            return None
        return int(n) if n.is_integer() else n
    text = clean(value).replace(" ", "")
    if not text:
        return None
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    else:
        text = text.replace(",", ".")
    try:
        n = float(text)
    except ValueError:
        return None
    if not math.isfinite(n):
        return None
    return int(n) if n.is_integer() else n


def load_rows(path: Path, sheet_name: str) -> list[dict[str, Any]]:
    wb = load_workbook(path)
    if sheet_name not in wb.sheetnames:
        raise SystemExit(f"Sheet not found: {sheet_name}. Available: {', '.join(wb.sheetnames)}")
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [clean(h) for h in rows[0]]
    out: list[dict[str, Any]] = []
    for raw in rows[1:]:
        row = {headers[i]: raw[i] if i < len(raw) else None for i in range(len(headers)) if headers[i]}
        if clean(row.get("id")):
            out.append(row)
    return out


def load_base_json(path: Path | None) -> dict[str, Any]:
    if path is None:
        return {"source": "motornet.it", "status": "ok", "schema": "cars_motornet_v1", "cars": []}
    payload = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(payload, dict) or not isinstance(payload.get("cars"), list):
        raise SystemExit("Base JSON must be an object with cars[]")
    return payload


def apply_row(car: dict[str, Any], row: dict[str, Any], clear_empty: bool) -> None:
    editable_fields = TEXT_FIELDS | NUMERIC_FIELDS
    unknown_fields = sorted(set(row.keys()) - set(CURATED_COLUMNS))
    if unknown_fields:
        # Unknown columns are intentionally ignored to avoid accidentally writing
        # helper spreadsheet data back into frontend JSON.
        pass

    for field in TEXT_FIELDS:
        if field not in row:
            continue
        value = clean(row.get(field))
        if value or clear_empty:
            car[field] = value

    for field in NUMERIC_FIELDS:
        if field not in row:
            continue
        value = parse_number(row.get(field))
        if value is not None:
            car[field] = value
        elif clear_empty:
            car.pop(field, None)

    # Keep manual notes for traceability only; not used by frontend calculations.
    note = clean(row.get("notes"))
    status = clean(row.get("status"))
    if note:
        car["_excel_note"] = note
    elif clear_empty:
        car.pop("_excel_note", None)
    if status:
        car["_excel_status"] = status
    elif clear_empty:
        car.pop("_excel_status", None)

    # Do not allow stale alternative consumption units to survive an explicit edit.
    if "consumption_l_100km" in row and parse_number(row.get("consumption_l_100km")) is not None:
        car.pop("consumption_kg_100km", None)
    if "consumption_kg_100km" in row and parse_number(row.get("consumption_kg_100km")) is not None:
        car.pop("consumption_l_100km", None)


def build_json(rows: list[dict[str, Any]], base_payload: dict[str, Any], clear_empty: bool) -> dict[str, Any]:
    payload = copy.deepcopy(base_payload)
    cars = payload.setdefault("cars", [])
    by_id = {clean(car.get("id")): car for car in cars if isinstance(car, dict) and clean(car.get("id"))}

    for row in rows:
        car_id = clean(row.get("id"))
        if not car_id:
            continue
        car = by_id.get(car_id)
        if car is None:
            car = {"id": car_id}
            cars.append(car)
            by_id[car_id] = car
        apply_row(car, row, clear_empty=clear_empty)

    payload["status"] = payload.get("status") or "ok"
    payload["schema"] = payload.get("schema") or "cars_motornet_v1"
    payload["_excel_roundtrip"] = True
    payload["_excel_curated_columns"] = CURATED_COLUMNS
    return payload


def main() -> int:
    parser = argparse.ArgumentParser(description="Convert cleaned Motornet Excel back to JSON.")
    parser.add_argument("--excel", required=True, help="Excel workbook exported by audit_motornet_quality.py")
    parser.add_argument("--sheet", default="catalog", help="Worksheet name")
    parser.add_argument("--base-json", default="data/cars_motornet.json", help="Original JSON used as base")
    parser.add_argument("--out", default="data/cars_motornet.cleaned.json", help="Output JSON path")
    parser.add_argument("--clear-empty", action="store_true", help="Allow empty Excel cells to clear JSON values")
    args = parser.parse_args()

    excel_path = Path(args.excel)
    base_path = Path(args.base_json) if args.base_json else None
    out_path = Path(args.out)

    if not excel_path.exists():
        raise SystemExit(f"Excel not found: {excel_path}")
    if base_path is not None and not base_path.exists():
        raise SystemExit(f"Base JSON not found: {base_path}")

    rows = load_rows(excel_path, args.sheet)
    payload = build_json(rows, load_base_json(base_path), clear_empty=args.clear_empty)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"Rows read: {len(rows)}")
    print(f"Cars written: {len(payload.get('cars', []))}")
    print(f"Output: {out_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
